import asyncio
import os
import json
import torch
from datetime import datetime
from TTS.api import TTS
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    filters,
    ContextTypes,
)

from openpyxl import Workbook, load_workbook

# ======================= إعداد التخزين =========================
USERS_JSON = "users.json"
USERS_XLSX = "users.xlsx"

def load_users():
    if os.path.exists(USERS_JSON):
        with open(USERS_JSON, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_users(users):
    with open(USERS_JSON, "w", encoding="utf-8") as f:
        json.dump(users, f, ensure_ascii=False, indent=2)

def add_user_to_excel(user_id, name):
    if os.path.exists(USERS_XLSX):
        wb = load_workbook(USERS_XLSX)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "Join Date"])

    ids = [str(cell.value) for cell in ws["A"]]
    if str(user_id) not in ids:
        ws.append([user_id, name, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        wb.save(USERS_XLSX)
        print(f"📄 تمت إضافة {name} إلى Excel")

        
        try:
            None
            
        except:
            pass  

def add_user_if_new(user_id, name):
    users = load_users()
    if str(user_id) not in users:
        users[str(user_id)] = {"name": name}
        save_users(users)
        add_user_to_excel(user_id, name)
        print(f"👤 مستخدم جديد: {name} ({user_id})")

# ======================= إعداد الصوت =========================
device = "cuda" if torch.cuda.is_available() else "cpu"
tts = TTS("tts_models/multilingual/multi-dataset/xtts_v2").to(device)

# ======================= الأوامر =========================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.message.from_user
    add_user_if_new(user.id, user.full_name)

    context.user_data["started"] = True
    await update.message.reply_text("👋 أرسل النص الذي تريد تحويله إلى صوت:")

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.message.from_user
    add_user_if_new(user.id, user.full_name)

    text = update.message.text
    context.user_data["text"] = text

    uid = update.message.from_user.id
    output_file = f"output_{uid}.wav"
    waiting_msg = await update.message.reply_text("⏳ جارٍ إنشاء الصوت...")

    try:
        tts.tts_to_file(
            text=text,
            speaker_wav="s1.wav",
            language="ar",
            file_path=output_file
        )
        await waiting_msg.delete()
        await update.message.reply_audio(audio=open(output_file, "rb"))

        keyboard = [[
            InlineKeyboardButton("-50", callback_data="rate_-50"),
            InlineKeyboardButton("-30", callback_data="rate_-30"),
            #InlineKeyboardButton("طبيعي", callback_data="rate_100")
        ]]
        await update.message.reply_text("اختر سرعة الصوت:", reply_markup=InlineKeyboardMarkup(keyboard))
    except Exception as e:
        await waiting_msg.delete()
        await update.message.reply_text(f"❌ خطأ أثناء إنشاء الصوت: {e}")

# ======================= سرعة الصوت =========================


async def handle_rate_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    text = context.user_data.get("text")

    
    try:
        None
        
    except:
        try:
            await query.answer()  
        except:
            pass  

    if not text:
        try:
            await query.edit_message_text("❌ لم يتم العثور على النص الأصلي.")
        except:
            pass
        return

    data = query.data
    uid = query.from_user.id

    # تحديد سرعة الصوت
    if data == "rate_-50":
        speed_value = 0.55
    elif data == "rate_-30":
        speed_value = 0.12
    else:
        speed_value = 1.0

    output_path = f"output_{uid}_{int(speed_value * 100)}.wav"
    print(output_path)

    try:
        await query.edit_message_text(f"🔄 يتم إعادة إنشاء الصوت بسرعة {int(speed_value * 100)}%...")
    except:
        pass  # إذا انتهت المهلة، تجاهل التعديل

    try:
        tts.tts_to_file(
            text=text,
            speaker_wav="speaker_sample.wav",
            language="ar",
            file_path=output_path,
            speed=speed_value
        )
        await query.message.reply_audio(audio=open(output_path, "rb"))
    except Exception as e:
        await query.message.reply_text(f"❌ خطأ أثناء إعادة توليد الصوت: {e}")



# ======================= تشغيل البوت =========================
async def main():
    while True:
        try:
            bot_token = "YOUR_TOKEN_HERE" 
            app = Application.builder().token(bot_token).build()

            app.add_handler(CommandHandler("start", start))
            app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
            app.add_handler(CallbackQueryHandler(handle_rate_choice))

            print("✅ البوت يعمل...")
            await app.initialize()
            await app.start()
            await app.updater.start_polling()
            await asyncio.Event().wait()
        except Exception as e:
            print(f"❌ حدث خطأ: {e}")
            print("🔁 إعادة تشغيل البوت خلال 2 ثوانٍ...")
            await asyncio.sleep(2)

# ✅ السطر المهم الذي يشغّل البوت
if __name__ == "__main__":
    asyncio.run(main())



